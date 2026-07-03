import openpyxl
import sys
import os
from logi.logi import logger
# ====================НАЧАЛО НАСТРОЙКИ ====================
EXCEL_FILE = "baze/best.xlsx"
SHEET_NAME = "title" # None = первый лист, или укажите имя листа

# Колонки (A=1, B=2, ...)
COL_INVENTORY = 1  # A - Инвентарный номер
COL_LOCATION = 2  # B - Пункт установки оборудования
COL_MUX = 4  # D - Мультиплекс
COL_TRANSMITTER = 7  # G - Наименование передатчика
# ====================КОНЕЦ НАСТРОЙКИ ====================

def load_excel_data(filepath):
    """Загружает данные из Excel файла"""
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        if SHEET_NAME:
            # точное имя листа
            ws = wb[SHEET_NAME]
        else:
            # последний сохраненный лист
            ws = wb.active
        # logger.info(f"Данные из файла:{filepath}===Лист:{ws.title}===Строк всего:{ws.max_row - 1}")

        # Создаём словарь: инвентарный номер -> (пункт установки, наименование)
        # можно добавлять значения и другие
        data = {}
        # начинаем со 2 строки
        for row in ws.iter_rows(min_row=2, values_only=True):
            inventory = row[COL_INVENTORY - 1]
            location = row[COL_LOCATION - 1]
            mux = row[COL_MUX - 1]
            transmitter = row[COL_TRANSMITTER - 1]

            if inventory:
                # Нормализуем ключ (убираем пробелы, приводим к верхнему регистру)
                key = str(inventory).strip().upper()
                data[key] = {
                    'location': location or "В-XL-МЕСТО-НЕ-УКАЗАНО",
                    'mux': mux or "В-XL-MUX-НЕ-УКАЗАНО",
                    'transmitter': transmitter or "В-XL-МОДЕЛЬ-НЕ-УКАЗАНО"
                }
        wb.close()
        return data

    except FileNotFoundError:
        logger.warning(f"Файл '{filepath}' не найден!")
        sys.exit(1)
    except Exception as e:
        logger.info(f"❌ Ошибка загрузки файла: {e}")
        sys.exit(1)


def search_inventory(data, query):
    """????????Ищет инвентарный номер по запросу"""
    query = query.strip().upper()
    # Точное совпадение
    if query in data:
        return data[query]
    # Если ничего не найдено — возвращаем словарь с заглушками
    return {
        'location': 'нет в XL',
        'mux': 'нет в XL',
        'transmitter': 'нет в XL'
    }
    # # Частичное совпадение  Зачем оно нужно вообше???
    # results = []
    # for key, value in data.items():
    #     if query in key:
    #         results.append(value)
    # return results

# ==================== ГЛАВНАЯ ФУНКЦИЯ ======================
def poisk_data(query):
    """ИТОГОВАЯ ФУНКЦИЯ ПОИСКА"""
    try:
        data = load_excel_data(EXCEL_FILE)
        results = search_inventory(data, query)
        return results
    except Exception as e:
        logger.warning(f"{query} не найден в XL!.Ошибка в read_xl/poisk_data() {e}")
# ==================== ГЛАВНАЯ ФУНКЦИЯ =======================
if __name__ == "__main__":
    clean_date = "05.03.2026"
    clean_inventory = "БАШ1880"
    # ===========================================
    results=poisk_data(query = clean_inventory)
    location = results['location']
    mux = "mux-"+str(results['mux'])
    transmitter = results['transmitter']
    new_filename = f"Протокол_{clean_date}_{location}_{mux}_{transmitter}_{clean_inventory}.pdf"
    print(new_filename)
    # print(location)
    # print(mux)
    # print(transmitter)
