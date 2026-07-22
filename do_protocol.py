import openpyxl
import sys
import os
from logi.logi import logger
from datetime import datetime, date
from input_data import input_rename, convert_xlsx_to_pdf
import random
from geo_coordinate import main_geo_coordinate, format_to_dms
from raschetni_azimut import main_raschetnii_azimut, simulate_measured_azimuth
from datetime import datetime, timedelta
import holidays

ru_holidays = holidays.Russia()

# ====================НАЧАЛО НАСТРОЙКИ ====================
# BASE_FILE = "baze/best.xlsx"
# BASE_SHEET_NAME = "title" # None = первый лист, или укажите имя листа
BASE_FILE = "baze/all_data_best.xlsx"
BASE_SHEET_NAME = "mux1"  # None = первый лист, или укажите имя листа

POINT_COLUMN = 4  # Колонка E (индекс 4, считая с 0) - "Пункт установки"
HEADER_ROW = 3  # Строка с основными заголовками (строка 4 в Excel, индекс 3)
DATA_START_ROW = 6  # Строка начала данных (строка 7 в Excel, индекс 6)

output_folder = 'posle'


# ====================КОНЕЦ НАСТРОЙКИ ====================
# ============НАЧАЛО ПОЛУЧЕНИЯ ДАННЫХ ИЗ all_data_best.xlsx ======


def format_cell_value(value):
    """Форматирует значение ячейки, конвертируя даты в строки формата DD.MM.YY"""
    # print(f"&&&&&&&&&&&&{value}")
    if value is None:
        return None

    # Если значение - дата или время
    if isinstance(value, (datetime, date)):
        # Форматируем как DD.MM.YY
        return value.strftime('%d.%m.%y')

    # Для всех остальных значений возвращаем как есть
    return value


def read_all_data(file_path, sheet_name=None):
    """Читает весь лист Excel и возвращает:
    - headers: список заголовков колонок
    - data: список строк (каждая строка - список значений)"""
    if not os.path.exists(file_path):
        logger.error(f"Файл не найден: {file_path}")
        return None, None
    try:
        # Убрали data_only=True чтобы корректно читать даты
        wb = openpyxl.load_workbook(file_path, read_only=True)
        if sheet_name is None:
            sheet = wb.active
        else:
            if sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
            else:
                logger.error(f"Лист '{sheet_name}' не найден. Доступны: {wb.sheetnames}")
                wb.close()
                return None, None

        # Читаем заголовки из указанной строки
        headers = []
        for cell in sheet[HEADER_ROW + 1]:  # +1 потому что openpyxl нумерует с 1
            headers.append(str(cell.value) if cell.value else f"Колонка_{cell.column}")

        # Читаем все данные начиная с DATA_START_ROW
        data = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=DATA_START_ROW + 1, values_only=True),
                                      start=DATA_START_ROW):
            # Пропускаем пустые строки
            if any(cell is not None for cell in row):
                # Форматируем каждую ячейку в строке (конвертируем даты)
                formatted_row = [format_cell_value(cell) for cell in row]
                data.append(formatted_row)

        wb.close()

        return headers, data

    except Exception as e:
        logger.error(f"Ошибка чтения файла: {e}")
        return None, None


def find_point_by_name(headers, data, point_name, point_column=POINT_COLUMN):
    """Ищет строку по названию пункта установки.
    Args:
        headers: список заголовков
        data: список строк
        point_name: название пункта (например, "БАЙМАК")
        point_column: индекс колонки с пунктами установки
    Returns:
        dict: словарь {название_колонки: значение} или None"""
    if not headers or not data:
        logger.error("Нет данных для поиска")
        return None

    # Нормализуем название для поиска (убираем пробелы, приводим к верхнему регистру)
    point_name_upper = point_name.strip().upper()
    # logger.info(f"Ищем пункт: '{point_name_upper}' в колонке {headers[point_column]}")
    for row_idx, row in enumerate(data, start=DATA_START_ROW + 1):
        if point_column < len(row):
            cell_value = row[point_column]
            if cell_value is not None:
                cell_value_str = str(cell_value).strip().upper()
                # Точное совпадение
                if cell_value_str == point_name_upper:
                    # logger.info(f"Найден пункт '{point_name}' в строке {row_idx}")
                    # Создаем словарь
                    result = {}
                    for col_idx, header in enumerate(headers):
                        if col_idx < len(row):
                            result[header] = row[col_idx]
                        else:
                            result[header] = None
                    return result

    logger.warning(f"Пункт '{point_name}' не найден")
    return None


def get_point_value(point_data, column_name):
    """Получает значение из данных пункта по имени колонки."""
    if not point_data:
        return None
    # Ищем колонку по имени (регистронезависимо)
    column_name_upper = column_name.upper()
    for key, value in point_data.items():
        if key and key.upper() == column_name_upper:
            return value
    # Если не нашли точное совпадение, ищем частичное
    for key, value in point_data.items():
        if key and column_name_upper in key.upper():
            return value
    return None


def get_next_workday(date_str: str) -> str:
    """
    Принимает дату в формате 'DD.MM.YYYY'.
    Возвращает дату следующего рабочего дня в том же формате.
    """
    date_obj = datetime.strptime(date_str, "%d.%m.%Y").date()

    while date_obj.weekday() >= 5 or date_obj in ru_holidays:
        date_obj += timedelta(days=1)

    return date_obj.strftime("%d.%m.%Y")


def main_itog_do_protocol(search_point, location_measure_metrics, date_protocol, naprazonnost_mediannaya_niz,
                          naprazonnost_mediannaya_verch):
    """ИТОГОВАЯ ФУНКЦИЯ"""
    # =======Начало входные данные=====
    itog_number = None
    # search_point = "Уфа"  # Место где стоит вышка
    # location_measure_metrics = "большетенькашево"  # Граничный населенный пункт где беруться замеры
    # date_protocol = "10.08.2016"
    # =======Конец входные данные======
    # 1. Читаем файл
    headers, data = read_all_data(BASE_FILE, BASE_SHEET_NAME)
    # 3. Ищем конкретный пункт
    logger.info(f"Поиск пункта: {search_point}")
    point_data = find_point_by_name(headers, data, search_point)
    print(point_data)

    # ПАРАМЕТРЫ
    # номер протокола
    if itog_number == None:
        itog_number = random.randint(100, 999)

    # район
    rayon = get_point_value(point_data, "Район")
    itog_rayon = search_point.strip().capitalize() + f"({rayon} р-н)"
    logger.info(f"Район: {itog_rayon}")

    # номер разрешения
    razrecshenie = get_point_value(point_data, "Разрешение на использование радиочастот")
    razrecshenie_data_vidachi = get_point_value(point_data, "Дата выдачим")
    itog_razrecshenie = f"Разрешение " + razrecshenie + " от " + str(razrecshenie_data_vidachi)
    logger.info(f"Разрешение: {itog_razrecshenie}")

    # координаты
    coordinates_shirota_1 = get_point_value(point_data, "Широта градусы")
    # if coordinates_shirota_1 is None:
    #     coordinates_shirota_1 = int(random.randint(10, 50))
    coordinates_shirota_2 = get_point_value(point_data, "Широта минуты")
    # if coordinates_shirota_2 is None:
    #     coordinates_shirota_2 = int(random.randint(10, 50))
    coordinates_shirota_3 = get_point_value(point_data, "Широта секунды")
    # if coordinates_shirota_3 is None:
    #     coordinates_shirota_3 = int(random.randint(10, 50))
    coordinates_dolgota_1 = get_point_value(point_data, "Долгота градусы")
    # if coordinates_dolgota_1 is None:
    #     coordinates_dolgota_1 = int(random.randint(10, 50))
    coordinates_dolgota_2 = get_point_value(point_data, "Долгота минуты")
    # if coordinates_dolgota_2 is None:
    #     coordinates_dolgota_2 = int(random.randint(10, 50))
    coordinates_dolgota_3 = get_point_value(point_data, "Долгота секунды")
    # if coordinates_dolgota_3 is None:
    #     coordinates_dolgota_3 = int(random.randint(10, 50))
    # Правильный формат: 55° 45' 20.88" N, 37° 37' 2.28" E
    itog_coordinates = f"{coordinates_shirota_1}°{coordinates_shirota_2}'{coordinates_shirota_3}\"N,{coordinates_dolgota_1}°{coordinates_dolgota_2}'{coordinates_dolgota_3}\"E"

    # мощность передатчика
    power = get_point_value(point_data, "Мощность, кВт")
    itog_power = f"{power} кВт"

    # высота подвеса
    height = get_point_value(point_data, "Высота подвеса над уровнем земли (из ЧТП)")
    itog_height = f"{height} м"

    # тип антены
    type_anten = get_point_value(point_data, "Тип антенны")
    itog_type_anten = f"{type_anten}"

    # коэффициент усиления
    koeff_ysilenia = get_point_value(point_data, "Ку , дБд")
    itog_koeff_ysilenia = f"{koeff_ysilenia} дБ"

    # средство измерений
    svidetelstvo_izmerenia = {"16.02.2027": ["17.02.2026", "С-БЕВ/17-02-2026/506326766"],
                              "20.01.2026": ["21.01.2025", "С-БЕВ/21-01-2025/404395610"],
                              "27.12.2024": ["28.12.2023", "С-БЕВ/28-12-2023/306877891"],
                              "08.09.2023": ["09.09.2022", "С-БЕВ/09-09-2022/187120735"],
                              "04.08.2022": ["05.08.2021", "С-БЕВ/05-08-2021/85940153"],
                              "12.07.2021": ["13.07.2020", "УС-20/430"],
                              "03.07.2020": ["04.07.2019", "УС-1864/19"],
                              "13.06.2019": ["14.06.2018", "8360/2018-717068"],
                              "21.06.2018": ["22.06.2017", "5057/2017-713547"],
                              "19.06.2017": ["20.06.2016", "2854/2016-710188"]}

    protocol_date = datetime.strptime(date_protocol, "%d.%m.%Y")
    # Ищем минимальный ключ, который больше date_protocol
    result_key = min((key for key in svidetelstvo_izmerenia if datetime.strptime(key, "%d.%m.%Y") > protocol_date),
                     key=lambda k: datetime.strptime(k, "%d.%m.%Y"))
    svidetelstvo_poverka_date = svidetelstvo_izmerenia[result_key][0]
    svidetelstvo_poverka_number = svidetelstvo_izmerenia[result_key][1]

    # Берем граничный населенный пункт и парсим координаты
    itog_location_measure_metrics = location_measure_metrics.strip().capitalize()
    point_geo_all = main_geo_coordinate(location_measure_metrics=itog_location_measure_metrics)
    print(point_geo_all)
    # format_to_dms(lat=point_geo_all[0][0], lon=point_geo_all[0][1])
    point_geo_1_shirota = format_to_dms(*point_geo_all[0])[0]
    point_geo_1_dolgota = format_to_dms(*point_geo_all[0])[1]
    point_geo_2_shirota = format_to_dms(*point_geo_all[1])[0]
    point_geo_2_dolgota = format_to_dms(*point_geo_all[1])[1]
    point_geo_3_shirota = format_to_dms(*point_geo_all[2])[0]
    point_geo_3_dolgota = format_to_dms(*point_geo_all[2])[1]
    point_geo_4_shirota = format_to_dms(*point_geo_all[3])[0]
    point_geo_4_dolgota = format_to_dms(*point_geo_all[3])[1]
    point_geo_5_shirota = format_to_dms(*point_geo_all[4])[0]
    point_geo_5_dolgota = format_to_dms(*point_geo_all[4])[1]
    point_geo_6_shirota = format_to_dms(*point_geo_all[5])[0]
    point_geo_6_dolgota = format_to_dms(*point_geo_all[5])[1]

    # Расчетный азимут
    point_1_two = f"{point_geo_1_shirota},{point_geo_1_dolgota}"
    azimut_raschetni_1 = round(main_raschetnii_azimut(p1=itog_coordinates, p2=point_1_two), 1)
    azimut_izmereni_1 = round(simulate_measured_azimuth(azimut_raschetni_1, sigma_deg=0.06), 1)
    point_2_two = f"{point_geo_2_shirota},{point_geo_2_dolgota}"
    azimut_raschetni_2 = round(main_raschetnii_azimut(p1=itog_coordinates, p2=point_2_two), 1)
    azimut_izmereni_2 = round(simulate_measured_azimuth(azimut_raschetni_2, sigma_deg=0.08), 1)
    point_3_two = f"{point_geo_3_shirota},{point_geo_3_dolgota}"
    azimut_raschetni_3 = round(main_raschetnii_azimut(p1=itog_coordinates, p2=point_3_two), 1)
    azimut_izmereni_3 = round(simulate_measured_azimuth(azimut_raschetni_3, sigma_deg=0.09), 1)
    point_4_two = f"{point_geo_4_shirota},{point_geo_4_dolgota}"
    azimut_raschetni_4 = round(main_raschetnii_azimut(p1=itog_coordinates, p2=point_4_two), 1)
    azimut_izmereni_4 = round(simulate_measured_azimuth(azimut_raschetni_4, sigma_deg=0.04), 1)
    point_5_two = f"{point_geo_5_shirota},{point_geo_5_dolgota}"
    azimut_raschetni_5 = round(main_raschetnii_azimut(p1=itog_coordinates, p2=point_5_two), 1)
    azimut_izmereni_5 = round(simulate_measured_azimuth(azimut_raschetni_5, sigma_deg=0.08), 1)
    point_6_two = f"{point_geo_6_shirota},{point_geo_6_dolgota}"
    azimut_raschetni_6 = round(main_raschetnii_azimut(p1=itog_coordinates, p2=point_6_two), 1)
    azimut_izmereni_6 = round(simulate_measured_azimuth(azimut_raschetni_6, sigma_deg=0.07), 1)

    # E норм и расч
    chenel_number = int(get_point_value(point_data, "ТВК"))

    data_chenel_number = {20: [466, 50.7, "EEEE"], 21: [474, 50.8, "EEEE"], 22: [482, 51.0, "EEEE"],
                          23: [490, 51.1, "27DB"], 24: [498, 51.3, "EEEE"], 25: [506, 51.4, "27DD"],
                          26: [514, 51.5, "27DE"], 27: [522, 51.7, "EEEE"], 28: [530, 51.8, "EEEE"],
                          29: [538, 51.9, "EEEE"], 30: [546, 52.1, "27E3"], 31: [554, 52.2, "27E5"],
                          32: [562, 52.3, "27DF"], 33: [570, 52.4, "27D9"], 34: [578, 52.6, "EEEE"],
                          35: [586, 52.7, "4EEC"], 36: [594, 52.8, "EEEE"], 37: [602, 52.9, "4EF0"],
                          38: [610, 53.0, "27E4"], 39: [618, 53.1, "EEEE"], 40: [626, 53.3, "EEEE"],
                          41: [634, 53.4, "27E1"], 42: [642, 53.5, "EEEE"], 43: [650, 53.6, "4EED"],
                          44: [658, 53.7, "4EEA"], 45: [666, 53.8, "27DC"], 46: [674, 53.9, "27DA"],
                          47: [682, 54.0, "EEEE"], 48: [690, 54.1, "EEEE"], 49: [698, 54.2, "EEEE"],
                          50: [706, 54.3, "EEEE"], 51: [714, 54.4, "EEEE"], 52: [722, 54.5, "EEEE"],
                          53: [730, 54.6, "EEEE"], 54: [738, 54.7, "EEEE"], 55: [746, 54.8, "EEEE"],
                          56: [754, 54.9, "4EE9"], 57: [762, 55.0, "4EF1"], 58: [770, 55.1, "4EEF"],
                          59: [778, 55.1, "EEEE"]}
    gauss_value = data_chenel_number[chenel_number][1]
    cell_id = data_chenel_number[chenel_number][2]
    naprazonnost_mediannaya_itog = round(random.randint(int(round(naprazonnost_mediannaya_niz, 2) * 100),
                                                        int(round(naprazonnost_mediannaya_verch, 2) * 100)) / 100, 2)
    if naprazonnost_mediannaya_itog <= gauss_value+1:
        naprazonnost_mediannaya_itog = gauss_value + 1.01

    # процент охвата населения
    prozent_ohvata_naselenia = round(get_point_value(point_data, "% охвата населения"), 2)
    itog_prozent_ohvata_naselenia = f"{prozent_ohvata_naselenia}"

    # запись
    input_rename(itog_rayon=itog_rayon, itog_razrecshenie=itog_razrecshenie, date_protocol=date_protocol,
                 itog_number=itog_number, itog_coordinates=itog_coordinates, itog_power=itog_power,
                 itog_height=itog_height, svidetelstvo_poverka_date=svidetelstvo_poverka_date,
                 svidetelstvo_poverka_number=svidetelstvo_poverka_number,
                 itog_location_measure_metrics=itog_location_measure_metrics, point_geo_1_shirota=point_geo_1_shirota,
                 point_geo_1_dolgota=point_geo_1_dolgota, point_geo_2_shirota=point_geo_2_shirota,
                 point_geo_2_dolgota=point_geo_2_dolgota, point_geo_3_shirota=point_geo_3_shirota,
                 point_geo_3_dolgota=point_geo_3_dolgota, point_geo_4_shirota=point_geo_4_shirota,
                 point_geo_4_dolgota=point_geo_4_dolgota, point_geo_5_shirota=point_geo_5_shirota,
                 point_geo_5_dolgota=point_geo_5_dolgota, point_geo_6_shirota=point_geo_6_shirota,
                 point_geo_6_dolgota=point_geo_6_dolgota, azimut_raschetni_1=azimut_raschetni_1,
                 azimut_raschetni_2=azimut_raschetni_2, azimut_raschetni_3=azimut_raschetni_3,
                 azimut_raschetni_4=azimut_raschetni_4, azimut_raschetni_5=azimut_raschetni_5,
                 azimut_raschetni_6=azimut_raschetni_6, azimut_izmereni_1=azimut_izmereni_1,
                 azimut_izmereni_2=azimut_izmereni_2, azimut_izmereni_3=azimut_izmereni_3,
                 azimut_izmereni_4=azimut_izmereni_4, azimut_izmereni_5=azimut_izmereni_5,
                 azimut_izmereni_6=azimut_izmereni_6, gauss_value=gauss_value, cell_id=cell_id,
                 itog_koeff_ysilenia=itog_koeff_ysilenia, itog_type_anten=itog_type_anten,
                 itog_prozent_ohvata_naselenia=itog_prozent_ohvata_naselenia,
                 naprazonnost_mediannaya_itog=naprazonnost_mediannaya_itog)


# ============КОНЕЦ ПОЛУЧЕНИЯ ДАННЫХ ИЗ all_data_best.xlsx ======
if __name__ == "__main__":
    all_sp = [["Уфа", "Байгильдино", "01.05.2016", 53.56, 58.56],
              ["Уфа", "Балтика", "02.06.2016", 58.32, 65.64],
              ["Уфа", "Охлебинино", "03.07.2016", 53.31, 55.85],
              ["Уфа", "Арсланово", "14.04.2016", 57.76, 57.80],
              ["Уфа", "Арово", "15.04.2016", 63.28, 63.50],
              ["Уфа", "Шарипово", "16.04.2016", 61.05, 61.11],
              ["Уфа", "Николаевка", "17.05.2016", 61.63, 62.49],
              ["Уфа", "Дмитриевка", "18.04.2016", 54.29, 54.46]]
    failed_objects = []
    for object in all_sp:
        try:
            search_point = object[0]
            location_measure_metrics = object[1]
            # date_protocol = object[2]
            date_protocol = get_next_workday(object[2])
            naprazonnost_mediannaya_niz = object[3]
            naprazonnost_mediannaya_verch = object[4]
            main_itog_do_protocol(search_point=search_point, location_measure_metrics=location_measure_metrics,
                                  date_protocol=date_protocol, naprazonnost_mediannaya_niz=naprazonnost_mediannaya_niz,
                                  naprazonnost_mediannaya_verch=naprazonnost_mediannaya_verch)
        except Exception as e:
            logger.error(f"Ошибка чтения файла: {e}")
            print(f"&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&   Не удалось сделать {object}&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&")
            # Сохраняем объект и текст ошибки
            failed_objects.append({"object": object, "error": str(e)})
    convert_xlsx_to_pdf(folder_name=output_folder)

    # Вывод красивого отчёта
    if failed_objects:
        print(f"\n❌ Итого не обработано: {len(failed_objects)}")
        for item in failed_objects:
            print(f"  • {item['object']} — {item['error']}")
    else:
        print("\n✅ Все объекты успешно обработаны!")
